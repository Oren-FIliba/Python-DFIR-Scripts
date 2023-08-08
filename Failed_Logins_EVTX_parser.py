import win32evtlog
import csv
import datetime
from collections import Counter
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image
import os

csv_file_path = rf'C:\Cases\{datetime.datetime.date(datetime.datetime.now())}_failed_logins.csv'
stats_file_path = rf'C:\Cases\{datetime.datetime.date(datetime.datetime.now())}_failed_login_statistics.csv'


def write_to_csv(login_info):
    """Write login information to the CSV file."""


    with open(csv_file_path, 'a', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        csv_writer.writerow(login_info)

def write_statistics_to_csv(user_stats, ip_stats, failure_reason_stats, sub_status_stats):
    """Write user and IP address statistics to the CSV file."""

    print("\nCreating Statistics....")

    with open(stats_file_path, 'w', newline='') as statsfile:
        csv_writer = csv.writer(statsfile)

        # Write user statistics
        csv_writer.writerow(['User', 'Failed Login Count'])
        for user, count in user_stats.most_common():
            csv_writer.writerow([user, count])

        csv_writer.writerow([])  # Add an empty row for separation

        # Write IP address statistics
        csv_writer.writerow(['IP Address', 'Failed Login Count'])
        for ip, count in ip_stats.most_common():
            csv_writer.writerow([ip, count])

        csv_writer.writerow([])  # Add an empty row for separation

        # Write additional statistics
        csv_writer.writerow(['Additional Statistics'])
        csv_writer.writerow(['User with Least Failed Logins', *user_stats.most_common()[-1]])
        csv_writer.writerow(['IP Least Seen', *ip_stats.most_common()[-1]])
        csv_writer.writerow(['Most Common Failure Reason', *failure_reason_stats.most_common(1)[0]])
        csv_writer.writerow(['Most Rare Failure Reason', *failure_reason_stats.most_common()[-1]])
        csv_writer.writerow(['Most Rare Sub Status', *sub_status_stats.most_common()[-1]])

        print(f"Completed!\n----------------------------------\n")

def generate_logs_per_day(csv_file_path, output_csv_file_path):
    # Read the CSV file into a DataFrame
    df = pd.read_csv(csv_file_path)

    # Convert the 'Date' column to datetime format
    df['Date'] = pd.to_datetime(df['Date'])

    # Group logs by date and count the occurrences
    logs_per_day = df.groupby(df['Date'].dt.date).size().reset_index(name='Count')

    # Write the results to a new CSV file
    logs_per_day.to_csv(output_csv_file_path, index=False)

    # Plot the graph
    plt.figure(figsize=(12, 6))
    plt.bar(logs_per_day['Date'], logs_per_day['Count'], color='skyblue')
    plt.xlabel('Date')
    plt.ylabel('Number of Logs')
    plt.title('Number of Logs Per Day')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(r'C:\Cases\logs_per_day.png')
    #plt.show()

def merge_csv_to_excel(csv1_path, csv2_path, csv3_path, excel_path, image_path):

    print("Making Logs Per Day graph .....")
    # Function to read a CSV file and return its rows as a list
    def read_csv(csv_path):
        with open(csv_path, 'r', newline='') as csv_file:
            reader = csv.reader(csv_file)
            return [row for row in reader]

    # Read the CSV files
    data1 = read_csv(csv1_path)
    data2 = read_csv(csv2_path)
    data3 = read_csv(csv3_path)

    # Create an Excel workbook
    workbook = openpyxl.Workbook()

    # Add each CSV data to a separate sheet in the Excel file
    workbook.create_sheet('Sheet1', index=0)
    workbook.create_sheet('Sheet2', index=1)
    sheet3 = workbook.create_sheet('Sheet3', index=2)

    # Write data from CSVs to each sheet
    for sheet, data in zip(workbook.worksheets, [data1, data2, data3]):
        for row in data:
            sheet.append(row)

    print(f"Completed!\n----------------------------------\n")
    print("Creating an output Excel File.....  ")

    # Delete the original CSV files
    os.remove(csv1_path)
    os.remove(csv2_path)
    os.remove(csv3_path)

    # Insert the image into Sheet3
    img = Image(image_path)
    sheet3.add_image(img, 'D1')

    # Save the Excel file
    workbook.save(excel_path)
    os.remove(image_path)
    print(f"Completed!\n----------------------------------\n\n\nExcel File (Containing 3 sheets) Path: {excel_path}\n")


if __name__ == "__main__":

    excel_path = fr'C:\Cases\{datetime.datetime.date(datetime.datetime.now())}_failed_logins.xlsx'
    image_path = r'C:\Cases\logs_per_day.png'
    lpd_output_csv_file_path = r'C:\Cases\logs_per_day.csv'
    server = 'localhost'
    logtype = 'Security'
    hand = win32evtlog.OpenEventLog(server, logtype)
    flags = win32evtlog.EVENTLOG_BACKWARDS_READ | win32evtlog.EVENTLOG_SEQUENTIAL_READ
    headers = ["Date", "Event Id", "User", "Host", "Logon_Type", "Method", "Failure_reason", "Status",
               "Sub_status", "Src_ip", "Src_port", "Log Source"]
    write_to_csv(headers)

    # Calculate the date and time for one day ago from the current time
    timeframe = input("timeframe: \n\n").strip()
    if "d" in timeframe:
        final_timeframe = datetime.datetime.now() - datetime.timedelta(days=int(timeframe[:-1]))
    if "h" in timeframe:
        final_timeframe = datetime.datetime.now() - datetime.timedelta(hours=int(timeframe[:-1]))
    if "m" in timeframe:
        final_timeframe = datetime.datetime.now() - datetime.timedelta(minutes=int(timeframe[:-1]))

    user_stats = Counter()
    ip_stats = Counter()
    failure_reason_stats = Counter()
    sub_status_stats = Counter()
    print(f"\nParsing Event Logs from [{final_timeframe}] - [{datetime.datetime.now()}] for Failed Logins .....")

    while True:
        events = win32evtlog.ReadEventLog(hand, flags, 0)
        if events:
            for event in events:
                if (
                    event.EventID == 4625
                    and event.TimeGenerated >= final_timeframe
                ):
                    date = str(event.TimeGenerated)
                    id_ = event.EventID
                    source = event.SourceName
                    user = event.StringInserts[5]
                    host = event.StringInserts[6]
                    logon_Type = event.StringInserts[10]
                    method = event.StringInserts[11]
                    failure_reason = event.StringInserts[8]
                    status = event.StringInserts[7]
                    sub_status = event.StringInserts[9]
                    src_ip = event.StringInserts[19]
                    src_port = event.StringInserts[20]

                    login_info = [date, id_, user, host, logon_Type,
                                  method, failure_reason, status, sub_status,
                                  src_ip, src_port, source]

                    write_to_csv(login_info)

                    # Count failed logins for each user, IP address, failure reason, and sub status
                    user_stats[user] += 1
                    ip_stats[src_ip] += 1
                    failure_reason_stats[failure_reason] += 1
                    sub_status_stats[sub_status] += 1

                if event.TimeGenerated < final_timeframe:
                    print(f"Completed!\n----------------------------------")
                    write_statistics_to_csv(user_stats, ip_stats, failure_reason_stats, sub_status_stats)
                    generate_logs_per_day(csv_file_path, lpd_output_csv_file_path)
                    merge_csv_to_excel(csv_file_path, stats_file_path, lpd_output_csv_file_path, excel_path, image_path)

                    exit()
